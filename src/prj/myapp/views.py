from django.shortcuts import render
from django.http import HttpResponse
from django.shortcuts import redirect
from import_export.formats.base_formats import CSV
import csv  # Bibliothèque standard de Python pour lire et écrire des fichiers CSV
from django.views import View
from django.http import HttpResponseRedirect
from .resources import WilayaResource
from .resources import MoughataaResource
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import load_workbook
import openpyxl
from .resources import CartProductResource
from django.http import JsonResponse
from django.core.paginator import Paginator


from django.views.generic import ListView, DetailView, CreateView, DeleteView, UpdateView
from django.urls import reverse_lazy
from .models.model import Wilaya, Moughataa, Commune, ProductType, Product, PointOfSale, ProductPrice , Cart, CartProduct
from django.db.models import Q  # Pour la recherche avancée



from django.shortcuts import render
import time

#def welcome_view(request):
 #   return render(request, "welcome.html")  # Affiche la page de bienvenue

#def home_view(request):
 #   return render(request, "dashboard.html")  # Charge la page d'accueil après 3s

def welcome_view(request):
    return render(request, "welcome.html")

@login_required
def home(request):
    return render(request, "home.html")  # ✅ Vérifie que le fichier home.html existe bien
# 📍 Liste des Wilayas
from django.core.paginator import Paginator

class WilayaListView(LoginRequiredMixin, ListView):
    model = Wilaya
    template_name = "myapp/wilaya_list.html"  # ✅ Modification ici
    context_object_name = "wilayas"
    paginate_by = 10  # 🔹 Nombre d'éléments par page

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")  # Récupère le terme de recherche
        if query:
            queryset = queryset.filter(
                Q(name__icontains=query) | Q(code__icontains=query)
            )  # Recherche par nom ou code
        return queryset

# 📍 Détails d'une Wilaya
class WilayaDetailView(LoginRequiredMixin,DetailView):
    model = Wilaya
    template_name = "myapp/wilaya_detail.html"  # ✅ Modification ici
    context_object_name = "wilaya"

# 📍 Ajouter une Wilaya
class WilayaCreateView(LoginRequiredMixin,CreateView):
    model = Wilaya
    fields = '__all__'
    template_name = "myapp/wilaya_form.html"  # ✅ Modification ici
    success_url = reverse_lazy('wilaya_list')

# 📍 Modifier une Wilaya
class WilayaUpdateView(LoginRequiredMixin,UpdateView):
    model = Wilaya
    fields = '__all__'
    template_name = "myapp/wilaya_form.html"  # ✅ Modification ici
    success_url = reverse_lazy('wilaya_list')

# 📍 Supprimer une Wilaya
class WilayaDeleteView(LoginRequiredMixin,DeleteView):
    model = Wilaya
    template_name = "myapp/wilaya_confirm_delete.html"  # ✅ Modification ici
    success_url = reverse_lazy('wilaya_list')

# 📍 Liste des Moughataas avec recherche
class MoughataaListView(LoginRequiredMixin, ListView):
    model = Moughataa
    template_name = "myapp/moughataa_list.html"
    context_object_name = "moughataas"
    paginate_by = 15  # ✅ Active la pagination (affiche 5 moughataas par page)

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")  # Récupère la recherche
        if query:
            queryset = queryset.filter(
                Q(label__icontains=query) | Q(code__icontains=query) | Q(wilaya__name__icontains=query)
            )
        return queryset
    
    

    
    
# 📍 Détails d'une Moughataa
class MoughataaDetailView(LoginRequiredMixin,DetailView):
    model = Moughataa
    template_name = "myapp/moughataa_detail.html"  # ✅ Modification ici
    context_object_name = "moughataa"
    
    

# 📍 Ajouter une Moughataa
class MoughataaCreateView(LoginRequiredMixin,CreateView):
    model = Moughataa
    fields = '__all__'
    template_name = "myapp/moughataa_form.html"  # ✅ Modification ici
    success_url = reverse_lazy('moughataa_list')

# 📍 Modifier une Moughataa
class MoughataaUpdateView(LoginRequiredMixin,UpdateView):
    model = Moughataa
    fields = '__all__'
    template_name = "myapp/moughataa_form.html"  # ✅ Modification ici
    success_url = reverse_lazy('moughataa_list')

# 📍 Supprimer une Moughataa
class MoughataaDeleteView(LoginRequiredMixin,DeleteView):
    model = Moughataa
    template_name = "myapp/moughataa_confirm_delete.html"  # ✅ Modification ici
    success_url = reverse_lazy('moughataa_list')

# 📍 Liste des Communes avec fonctionnalité de recherche
class CommuneListView(LoginRequiredMixin, ListView):
    model = Commune
    template_name = "myapp/commune_list.html"
    context_object_name = "communes"
    paginate_by = 20  # ✅ Active la pagination (affiche 5 communes par page)

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")  # Récupère la recherche
        if query:
            queryset = queryset.filter(
                Q(nom__icontains=query) | Q(pk__icontains=query) | Q(moughataa__label__icontains=query)
            )
        return queryset



# 📍 Détails d'une Commune
class CommuneDetailView(LoginRequiredMixin,DetailView):
    model = Commune
    template_name = "myapp/commune_detail.html"
    context_object_name = "commune"

# 📍 Ajouter une Commune
class CommuneCreateView(LoginRequiredMixin,CreateView):
    model = Commune
    fields = '__all__'
    template_name = "myapp/commune_form.html"
    success_url = reverse_lazy('commune_list')

# 📍 Modifier une Commune
class CommuneUpdateView(LoginRequiredMixin,UpdateView):
    model = Commune
    fields = '__all__'
    template_name = "myapp/commune_form.html"
    success_url = reverse_lazy('commune_list')

# 📍 Supprimer une Commune
class CommuneDeleteView(LoginRequiredMixin,DeleteView):
    model = Commune
    template_name = "myapp/commune_confirm_delete.html"
    success_url = reverse_lazy('commune_list')
    
    
# 📍 Liste des Types de Produits avec Recherche et Suppression en masse
class ProductTypeListView(LoginRequiredMixin, ListView):
    model = ProductType
    template_name = "myapp/producttype_list.html"
    context_object_name = "producttypes"
    paginate_by = 10  # ✅ Active la pagination

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")  # ✅ Récupère la valeur du champ de recherche

        if query:
            queryset = queryset.filter(
                Q(nom__icontains=query) |
                Q(description__icontains=query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["query"] = self.request.GET.get("q", "")  # ✅ Ajoute la requête au contexte
        return context


# 📍 Détails d'un Type de Produit
class ProductTypeDetailView(LoginRequiredMixin,DetailView):
    model = ProductType
    template_name = "myapp/producttype_detail.html"
    context_object_name = "producttype"

# 📍 Ajouter un Type de Produit
class ProductTypeCreateView(LoginRequiredMixin,CreateView):
    model = ProductType
    fields = '__all__'
    template_name = "myapp/producttype_form.html"
    success_url = reverse_lazy('producttype_list')

# 📍 Modifier un Type de Produit
class ProductTypeUpdateView(LoginRequiredMixin,UpdateView):
    model = ProductType
    fields = '__all__'
    template_name = "myapp/producttype_form.html"
    success_url = reverse_lazy('producttype_list')

# 📍 Supprimer un Type de Produit
class ProductTypeDeleteView(LoginRequiredMixin,DeleteView):
    model = ProductType
    template_name = "myapp/producttype_confirm_delete.html"
    success_url = reverse_lazy('producttype_list')

# 📍 Liste des Produits
class ProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = "myapp/product_list.html"
    context_object_name = "products"
    paginate_by = 10  # ✅ Active la pagination

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")  # ✅ Récupère la valeur du champ de recherche

        if query:
            queryset = queryset.filter(
                Q(nom__icontains=query) |
                Q(unite_mesure__icontains=query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["query"] = self.request.GET.get("q", "")  # ✅ Ajoute la requête au contexte
        return context

# 📍 Détails d'un Produit
class ProductDetailView(LoginRequiredMixin,DetailView):
    model = Product
    template_name = "myapp/product_detail.html"
    context_object_name = "product"

# 📍 Ajouter un Produit
class ProductCreateView(LoginRequiredMixin,CreateView):
    model = Product
    fields = '__all__'
    template_name = "myapp/product_form.html"
    success_url = reverse_lazy('product_list')

# 📍 Modifier un Produit
class ProductUpdateView(LoginRequiredMixin,UpdateView):
    model = Product
    fields = '__all__'
    template_name = "myapp/product_form.html"
    success_url = reverse_lazy('product_list')

# 📍 Supprimer un Produit
class ProductDeleteView(LoginRequiredMixin,DeleteView):
    model = Product
    template_name = "myapp/product_confirm_delete.html"
    success_url = reverse_lazy('product_list')

# 📍 Liste des Points de Vente avec suppression multiple
class PointOfSaleListView(LoginRequiredMixin, ListView):
    model = PointOfSale
    template_name = "myapp/pointofsale_list.html"
    context_object_name = "pointofsales"
    paginate_by = 15  # ✅ Active la pagination (affiche 5 points de vente par page)

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")  # Récupère la recherche
        if query:
            queryset = queryset.filter(
                Q(nom__icontains=query) | Q(type__icontains=query) | Q(commune__nom__icontains=query)
            )
        return queryset


# 📍 Détails d'un Point de Vente
class PointOfSaleDetailView(LoginRequiredMixin,DetailView):
    model = PointOfSale
    template_name = "myapp/pointofsale_detail.html"
    context_object_name = "pointofsale"

# 📍 Ajouter un Point de Vente
class PointOfSaleCreateView(LoginRequiredMixin,CreateView):
    model = PointOfSale
    fields = '__all__'
    template_name = "myapp/pointofsale_form.html"
    success_url = reverse_lazy('pointofsale_list')

# 📍 Modifier un Point de Vente
class PointOfSaleUpdateView(LoginRequiredMixin,UpdateView):
    model = PointOfSale
    fields = '__all__'
    template_name = "myapp/pointofsale_form.html"
    success_url = reverse_lazy('pointofsale_list')

# 📍 Supprimer un Point de Vente
class PointOfSaleDeleteView(LoginRequiredMixin,DeleteView):
    model = PointOfSale
    template_name = "myapp/pointofsale_confirm_delete.html"
    success_url = reverse_lazy('pointofsale_list')

# 📍 Liste des Prix des Produits
class ProductPriceListView(LoginRequiredMixin, ListView):
    model = ProductPrice
    template_name = "myapp/productprice_list.html"
    context_object_name = "productprices"
    paginate_by = 10  # ✅ Active la pagination

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")  # ✅ Récupère la valeur du champ de recherche

        if query:
            queryset = queryset.filter(
                Q(produit__nom__icontains=query) |
                Q(point_vente__nom__icontains=query) |
                Q(valeur__icontains=query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["query"] = self.request.GET.get("q", "")  # ✅ Ajoute la requête au contexte
        return context
# 📍 Détails d'un Prix de Produit
class ProductPriceDetailView(LoginRequiredMixin,DetailView):
    model = ProductPrice
    template_name = "myapp/productprice_detail.html"
    context_object_name = "productprice"

# 📍 Ajouter un Prix de Produit
class ProductPriceCreateView(LoginRequiredMixin,CreateView):
    model = ProductPrice
    fields = '__all__'
    template_name = "myapp/productprice_form.html"
    success_url = reverse_lazy('productprice_list')

# 📍 Modifier un Prix de Produit
class ProductPriceUpdateView(LoginRequiredMixin,UpdateView):
    model = ProductPrice
    fields = '__all__'
    template_name = "myapp/productprice_form.html"
    success_url = reverse_lazy('productprice_list')

# 📍 Supprimer un Prix de Produit
class ProductPriceDeleteView(LoginRequiredMixin,DeleteView):
    model = ProductPrice
    template_name = "myapp/productprice_confirm_delete.html"
    success_url = reverse_lazy('productprice_list')


# 📍 Liste des Paniers
class CartListView(LoginRequiredMixin, ListView):
    model = Cart
    template_name = "myapp/cart_list.html"
    context_object_name = "carts"
    paginate_by = 10  # ✅ Active la pagination

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")  # ✅ Récupère la valeur du champ de recherche

        if query:
            queryset = queryset.filter(
                Q(nom__icontains=query) |
                Q(description__icontains=query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["query"] = self.request.GET.get("q", "")  # ✅ Ajoute la requête au contexte
        return context

# 📍 Détails d'un Panier
class CartDetailView(LoginRequiredMixin,DetailView):
    model = Cart
    template_name = "myapp/cart_detail.html"
    context_object_name = "cart"

# 📍 Ajouter un Panier
class CartCreateView(LoginRequiredMixin,CreateView):
    model = Cart
    fields = '__all__'
    template_name = "myapp/cart_form.html"
    success_url = reverse_lazy('cart_list')

# 📍 Modifier un Panier
class CartUpdateView(LoginRequiredMixin,UpdateView):
    model = Cart
    fields = '__all__'
    template_name = "myapp/cart_form.html"
    success_url = reverse_lazy('cart_list')

# 📍 Supprimer un Panier
class CartDeleteView(LoginRequiredMixin,DeleteView):
    model = Cart
    template_name = "myapp/cart_confirm_delete.html"
    success_url = reverse_lazy('cart_list')

# 📍 Liste des Produits dans un Panier
class CartProductListView(LoginRequiredMixin, ListView):
    model = CartProduct
    template_name = "myapp/cartproduct_list.html"
    context_object_name = "cartproducts"
    paginate_by = 15  # ✅ Active la pagination

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get("q")  # ✅ Récupère la valeur du champ de recherche

        if query:
            queryset = queryset.filter(
                Q(produit__nom__icontains=query) |
                Q(panier__nom__icontains=query) |
                Q(poids__icontains=query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["query"] = self.request.GET.get("q", "")  # ✅ Ajoute la requête au contexte
        return context

    

# 📍 Détails d'un Produit dans un Panier
class CartProductDetailView(LoginRequiredMixin,DetailView):
    model = CartProduct
    template_name = "myapp/cartproduct_detail.html"
    context_object_name = "cartproduct"
# 📍 Ajouter un Produit dans un Panier
class CartProductCreateView(LoginRequiredMixin,CreateView):
    model = CartProduct
    fields = '__all__'
    template_name = "myapp/cartproduct_form.html"
    success_url = reverse_lazy('cartproduct_list')

# 📍 Modifier un Produit dans un Panier
class CartProductUpdateView(LoginRequiredMixin,UpdateView):
    model = CartProduct
    fields = '__all__'
    template_name = "myapp/cartproduct_form.html"
    success_url = reverse_lazy('cartproduct_list')

# 📍 Supprimer un Produit dans un Panier
class CartProductDeleteView(LoginRequiredMixin,DeleteView):
    model = CartProduct
    template_name = "myapp/cartproduct_confirm_delete.html"
    success_url = reverse_lazy('cartproduct_list')




# Exporter les données au format CSV
class WilayaExportView(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        resource = WilayaResource()
        dataset = resource.export()

        # Déterminer le format d'exportation (CSV ou Excel) à partir d'un paramètre GET
        export_format = request.GET.get("format", "csv").lower()

        if export_format == "xlsx":
            response = HttpResponse(dataset.xlsx, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename="wilayas.xlsx"'
        else:  # Par défaut, exporter en CSV
            response = HttpResponse(dataset.csv, content_type="text/csv")
            response['Content-Disposition'] = 'attachment; filename="wilayas.csv"'

        return response


class MoughataaExportView(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        resource = MoughataaResource()
        dataset = resource.export()

        # Déterminer le format d'exportation (CSV ou Excel) à partir d'un paramètre GET
        export_format = request.GET.get("format", "csv").lower()

        if export_format == "xlsx":
            response = HttpResponse(
                dataset.xlsx,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = 'attachment; filename="moughataas.xlsx"'
        else:  # Par défaut, exporter en CSV
            response = HttpResponse(dataset.csv, content_type="text/csv")
            response['Content-Disposition'] = 'attachment; filename="moughataas.csv"'

        return response


from myapp.resources import CommuneResource

class CommuneExportView(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        resource = CommuneResource()
        dataset = resource.export()

        # Vérifier le format d'exportation demandé (CSV ou Excel)
        export_format = request.GET.get("format", "csv").lower()

        if export_format == "xlsx":
            response = HttpResponse(
                dataset.xlsx,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = 'attachment; filename="communes.xlsx"'
        else:  # Par défaut, exporter en CSV
            response = HttpResponse(dataset.csv, content_type="text/csv")
            response['Content-Disposition'] = 'attachment; filename="communes.csv"'

        return response

class PointOfSaleExportView(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        resource = PointOfSaleResource()  # Remplacez par votre resource pour PointOfSale
        dataset = resource.export()

        # Déterminer le format d'exportation (CSV ou Excel) à partir d'un paramètre GET
        export_format = request.GET.get("format", "csv").lower()

        if export_format == "xlsx":
            response = HttpResponse(
                dataset.xlsx,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = 'attachment; filename="point_of_sales.xlsx"'
        else:  # Par défaut, exporter en CSV
            response = HttpResponse(dataset.csv, content_type="text/csv")
            response['Content-Disposition'] = 'attachment; filename="point_of_sales.csv"'

        return response


from .resources import ProductTypeResource

class ProductTypeExportView(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        resource = ProductTypeResource()  # Classe resource pour ProductType
        dataset = resource.export()

        # Export en Excel
        response = HttpResponse(
            dataset.xlsx,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="product_types.xlsx"'
        return response
    
class ProductExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        resource = ProductResource()  # Classe resource pour Product
        dataset = resource.export()

        response = HttpResponse(
            dataset.xlsx,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        return response


from .resources import ProductPriceResource

class ProductPriceExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        resource = ProductPriceResource()  # Utilisation de la classe resource
        dataset = resource.export()

        response = HttpResponse(
            dataset.xlsx,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="product_prices.xlsx"'
        return response
    
from django.http import HttpResponse
from .resources import CartResource

class CartExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        resource = CartResource()  # Utilisation de la classe resource
        dataset = resource.export()

        response = HttpResponse(
            dataset.xlsx,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="carts.xlsx"'
        return response


# 📤 Exporter CartProduct en Excel
class CartProductExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        resource = CartProductResource()
        dataset = resource.export()
        response = HttpResponse(dataset.xlsx, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="cartproducts.xlsx"'
        return response


from django.shortcuts import render
from django.http import HttpResponseRedirect
from myapp.resources import WilayaResource
from tablib import Dataset
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator


class WilayaImportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Affiche le formulaire pour importer un fichier
        return render(request, "myapp/wilaya_import.html")

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return render(request, "myapp/wilaya_import.html", {"errors": ["Aucun fichier n'a été fourni."]})

        try:
            # Vérifier l'extension du fichier
            if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                # Charger le fichier Excel avec openpyxl
                workbook = load_workbook(file)
                sheet = workbook.active  # Récupère la première feuille
                data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Sauter l'en-tête
                    if row[0] and row[1]:  # Assurez-vous que les colonnes 'code' et 'name' existent
                        data.append([row[0], row[1]])  # Ajoute 'code' et 'name' à la liste
            else:
                return render(request, "myapp/wilaya_import.html", {"errors": ["Le fichier doit être au format Excel (.xlsx ou .xls)."]})
        except Exception as e:
            return render(request, "myapp/wilaya_import.html", {"errors": [f"Erreur lors de la lecture du fichier Excel : {str(e)}"]})

        # Préparer le dataset pour l'importation
        resource = WilayaResource()
        dataset = Dataset()
        dataset.headers = ['code', 'name']  # Définir les colonnes attendues
        for row in data:
            dataset.append(row)  # Ajouter les lignes extraites du fichier Excel

        try:
            # Simulation d'importation avec dry_run=True
            result = resource.import_data(dataset, dry_run=True)
            if not result.has_errors():
                resource.import_data(dataset, dry_run=False)  # Import réel
                return HttpResponseRedirect('/wilayas/')
            else:
                return render(request, "myapp/wilaya_import.html", {"errors": result.row_errors})
        except Exception as e:
            return render(request, "myapp/wilaya_import.html", {"errors": [str(e)]})




class MoughataaImportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Affiche le formulaire pour importer un fichier
        return render(request, "myapp/moughataa_import.html")

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return render(request, "myapp/moughataa_import.html", {"errors": ["Aucun fichier n'a été fourni."]})

        try:
            # Vérification de l'extension du fichier
            if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                # Lecture du fichier Excel
                workbook = load_workbook(file)
                sheet = workbook.active  # Récupère la première feuille
                data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Sauter l'en-tête
                    if row[0] and row[1] and row[2]:  # Vérifie que les colonnes nécessaires existent
                        data.append([row[0], row[1], row[2]])  # 'code', 'label', 'wilaya'
            else:
                return render(request, "myapp/moughataa_import.html", {"errors": ["Le fichier doit être au format Excel (.xlsx ou .xls)."]})
        except Exception as e:
            return render(request, "myapp/moughataa_import.html", {"errors": [f"Erreur lors de la lecture du fichier Excel : {str(e)}"]})

        # Préparation des données pour l'importation
        resource = MoughataaResource()
        dataset = Dataset()
        dataset.headers = ['code', 'label', 'wilaya']  # Définir les colonnes attendues
        for row in data:
            try:
                # Trouver l'instance de Wilaya basée sur son nom
                wilaya = Wilaya.objects.get(name=row[2])
                # Ajouter les données au dataset
                dataset.append([row[0], row[1], wilaya.pk])
            except Wilaya.DoesNotExist:
                return render(request, "myapp/moughataa_import.html", {
                    "errors": [f"La Wilaya '{row[2]}' n'existe pas dans la base de données."]
                })

        try:
            # Simulation d'importation avec dry_run=True
            result = resource.import_data(dataset, dry_run=True)
            if not result.has_errors():
                # Import réel si pas d'erreurs
                resource.import_data(dataset, dry_run=False)
                return HttpResponseRedirect('/moughataas/')
            else:
                return render(request, "myapp/moughataa_import.html", {"errors": result.row_errors})
        except Exception as e:
            return render(request, "myapp/moughataa_import.html", {"errors": [str(e)]})



from myapp.resources import CommuneResource



class CommuneImportView(LoginRequiredMixin , View):
    def get(self, request, *args, **kwargs):
        # Affiche le formulaire pour importer un fichier
        return render(request, "myapp/commune_import.html")

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return render(request, "myapp/commune_import.html", {"errors": ["Aucun fichier n'a été fourni."]})

        # Vérifier si le fichier est un fichier Excel
        if not (file.name.endswith('.xls') or file.name.endswith('.xlsx')):
            return render(request, "myapp/commune_import.html", {"errors": ["Seuls les fichiers Excel sont acceptés."]})

        try:
            # Lecture du fichier Excel
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active  # Obtenir la première feuille du fichier

            # Vérification des colonnes attendues
            required_columns = ['nom', 'moughataa']
            detected_columns = [cell.value for cell in sheet[1]]  # Lire les en-têtes de la première ligne
            missing_cols = [col for col in required_columns if col not in detected_columns]

            if missing_cols:
                return render(request, "myapp/commune_import.html", {
                    "errors": [
                        f"Le fichier Excel est invalide. Colonnes manquantes : {', '.join(missing_cols)}",
                        f"Colonnes détectées : {', '.join(detected_columns or [])}"
                    ]
                })

            # Préparation des données pour l'importation
            resource = CommuneResource()
            dataset = Dataset()
            dataset.headers = required_columns

            # Lire les données à partir de la deuxième ligne
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    nom, moughataa_label = row
                    # Trouver l'instance de Moughataa basée sur son label
                    moughataa = Moughataa.objects.get(label=moughataa_label)
                    dataset.append([nom, moughataa.pk])
                except Moughataa.DoesNotExist:
                    return render(request, "myapp/commune_import.html", {
                        "errors": [f"La Moughataa '{moughataa_label}' n'existe pas dans la base de données."]
                    })

            # Simulation d'importation avec dry_run=True
            result = resource.import_data(dataset, dry_run=True)
            if not result.has_errors():
                # Import réel si pas d'erreurs
                resource.import_data(dataset, dry_run=False)
                return HttpResponseRedirect('/communes/')
            else:
                return render(request, "myapp/commune_import.html", {"errors": result.row_errors})

        except Exception as e:
            return render(request, "myapp/commune_import.html", {"errors": [f"Erreur lors de la lecture du fichier : {str(e)}"]})



from django.http import HttpResponseRedirect
from .resources import PointOfSaleResource
from .resources import ProductResource

from django.shortcuts import render
from django.http import HttpResponseRedirect
from openpyxl import load_workbook
from myapp.resources import ProductResource
from myapp.models import ProductType
from tablib import Dataset

from django.shortcuts import render
from django.http import HttpResponseRedirect
from openpyxl import load_workbook
from tablib import Dataset
from myapp.resources import ProductResource
from myapp.models import ProductType, Product

from django.shortcuts import render
from django.http import HttpResponseRedirect
from openpyxl import load_workbook
from tablib import Dataset
from myapp.models import ProductType
from myapp.resources import ProductResource

class ProductImportView(LoginRequiredMixin , View):
    def get(self, request, *args, **kwargs):
        return render(request, "myapp/product_import.html")

    def post(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return render(request, "myapp/product_import.html", {"errors": ["Aucun fichier n'a été fourni."]})

        # Vérifier l'extension
        if not (file.name.endswith('.xls') or file.name.endswith('.xlsx')):
            return render(request, "myapp/product_import.html", {"errors": ["Seuls les fichiers Excel sont acceptés."]})

        try:
            # Charger le fichier Excel
            workbook = load_workbook(file)
            sheet = workbook.active

            # Mapping des types de produit (nom -> ID)
            product_type_mapping = {pt.nom: pt.id for pt in ProductType.objects.all()}

            dataset = Dataset()
            dataset.headers = ['nom', 'description', 'type_produit', 'unite_mesure']

            errors = []
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    nom, description, type_produit_nom, unite_mesure = row

                    # Valider et mapper le type de produit
                    if type_produit_nom not in product_type_mapping:
                        errors.append(f"Ligne {row_index}: Type de produit '{type_produit_nom}' inconnu.")
                        continue

                    type_produit_id = product_type_mapping[type_produit_nom]
                    dataset.append([nom, description, type_produit_id, unite_mesure])

                except Exception as e:
                    errors.append(f"Ligne {row_index}: Erreur - {str(e)}")

            if errors:
                return render(request, "myapp/product_import.html", {"errors": errors})

            # Importation des données
            resource = ProductResource()
            result = resource.import_data(dataset, dry_run=True)

            if result.has_errors():
                return render(request, "myapp/product_import.html", {"errors": result.row_errors})

            resource.import_data(dataset, dry_run=False)
            return HttpResponseRedirect("/products/")

        except Exception as e:
            return render(request, "myapp/product_import.html", {"errors": [f"Erreur lors de la lecture du fichier : {str(e)}"]})


from django.shortcuts import render
from django.http import HttpResponseRedirect
from openpyxl import load_workbook
from tablib import Dataset
from myapp.resources import PointOfSaleResource
from myapp.models import Commune
from django.views import View

class PointOfSaleImportView(LoginRequiredMixin , View):
    def get(self, request, *args, **kwargs):
        # Affiche le formulaire pour importer un fichier
        return render(request, "myapp/pointofsale_import.html")

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return render(request, "myapp/pointofsale_import.html", {"errors": ["Aucun fichier n'a été fourni."]})

        # Vérifier si le fichier est un fichier Excel
        if not (file.name.endswith('.xls') or file.name.endswith('.xlsx')):
            return render(request, "myapp/pointofsale_import.html", {"errors": ["Seuls les fichiers Excel sont acceptés."]})

        try:
            # Lecture du fichier Excel
            workbook = load_workbook(file)
            sheet = workbook.active  # Obtenir la première feuille du fichier

            # Préparation des données pour l'importation
            resource = PointOfSaleResource()
            dataset = Dataset()
            dataset.headers = ['nom', 'type', 'commune', 'gps_lat', 'gps_lon']  # Colonnes attendues

            row_num = 1  # Pour traquer la ligne en cas d'erreur
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorer l'en-tête
                row_num += 1
                try:
                    # Extraire les données de chaque ligne
                    nom, type, commune_name, gps_lat, gps_lon = row

                    # Vérifier si la commune existe
                    commune = Commune.objects.get(nom=commune_name)

                    # Ajouter les données au dataset
                    dataset.append([nom, type, commune.pk, gps_lat, gps_lon])
                except Commune.DoesNotExist:
                    return render(request, "myapp/pointofsale_import.html", {
                        "errors": [f"Ligne {row_num}: La commune '{commune_name}' n'existe pas dans la base de données."]
                    })
                except ValueError as e:
                    return render(request, "myapp/pointofsale_import.html", {
                        "errors": [f"Ligne {row_num}: Erreur - {str(e)}"]
                    })

            # Simulation d'importation avec dry_run=True
            result = resource.import_data(dataset, dry_run=True)
            if not result.has_errors():
                # Import réel si pas d'erreurs
                resource.import_data(dataset, dry_run=False)
                return HttpResponseRedirect('/pointofsales/')
            else:
                return render(request, "myapp/pointofsale_import.html", {"errors": result.row_errors})

        except Exception as e:
            return render(request, "myapp/pointofsale_import.html", {
                "errors": [f"Erreur lors de la lecture du fichier Excel : {str(e)}"]
            })


class ProductTypeImportView(LoginRequiredMixin , View):
    def get(self, request, *args, **kwargs):
        # Affiche le formulaire pour importer un fichier
        return render(request, "myapp/producttype_import.html")

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return render(request, "myapp/producttype_import.html", {"errors": ["Aucun fichier n'a été fourni."]})

        # Vérifier si le fichier est un fichier Excel
        if not (file.name.endswith('.xls') or file.name.endswith('.xlsx')):
            return render(request, "myapp/producttype_import.html", {"errors": ["Seuls les fichiers Excel sont acceptés."]})

        try:
            # Lecture du fichier Excel
            workbook = load_workbook(file)
            sheet = workbook.active  # Obtenir la première feuille du fichier

            # Préparation des données pour l'importation
            resource = ProductTypeResource()
            dataset = Dataset()
            dataset.headers = ['nom', 'description']  # Colonnes attendues

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(row):  # Vérifie que toutes les colonnes sont remplies
                    dataset.append(row)

            # Simulation d'importation avec dry_run=True
            result = resource.import_data(dataset, dry_run=True)
            if not result.has_errors():
                # Import réel si pas d'erreurs
                resource.import_data(dataset, dry_run=False)
                return HttpResponseRedirect('/producttypes/')
            else:
                return render(request, "myapp/producttype_import.html", {"errors": result.row_errors})

        except Exception as e:
            return render(request, "myapp/producttype_import.html", {"errors": [f"Erreur lors de la lecture du fichier Excel : {str(e)}"]})




from django.http import HttpResponseRedirect
from .resources import ProductPriceResource
class ProductPriceImportView(View):
    def get(self, request, *args, **kwargs):
        return render(request, "myapp/productprice_import.html")

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return render(request, "myapp/productprice_import.html", {"errors": ["Aucun fichier n'a été fourni."]})

        if not (file.name.endswith('.xls') or file.name.endswith('.xlsx')):
            return render(request, "myapp/productprice_import.html", {"errors": ["Seuls les fichiers Excel sont acceptés."]})

        try:
            workbook = load_workbook(file)
            sheet = workbook.active

            dataset = Dataset()
            dataset.headers = ['produit', 'point_vente', 'valeur', 'date_from', 'date_to']  # Colonnes attendues

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    produit_nom, point_vente_nom, valeur, date_from, date_to = row

                    # Vérifier si le produit existe
                    produits = Product.objects.filter(nom=produit_nom)
                    if produits.exists():
                        produit = produits.first()  # Sélectionner le premier produit trouvé
                    else:
                        return render(request, "myapp/productprice_import.html", {
                            "errors": [f"Le produit '{produit_nom}' n'existe pas."]
                        })

                    # Vérifier si le point de vente existe
                    points_vente = PointOfSale.objects.filter(nom=point_vente_nom)
                    if points_vente.exists():
                        point_vente = points_vente.first()  # Sélectionner le premier point de vente trouvé
                    else:
                        return render(request, "myapp/productprice_import.html", {
                            "errors": [f"Le point de vente '{point_vente_nom}' n'existe pas."]
                        })

                    # Ajouter la ligne au dataset
                    dataset.append([produit.id, point_vente.id, valeur, date_from, date_to])

                except Exception as e:
                    return render(request, "myapp/productprice_import.html", {
                        "errors": [f"Erreur lors du traitement de la ligne : {str(e)}"]
                    })

            # Simulation d'importation avec dry_run=True
            resource = ProductPriceResource()
            result = resource.import_data(dataset, dry_run=True)
            if not result.has_errors():
                resource.import_data(dataset, dry_run=False)
                return HttpResponseRedirect('/productprices/')
            else:
                return render(request, "myapp/productprice_import.html", {"errors": result.row_errors})

        except Exception as e:
            return render(request, "myapp/productprice_import.html", {"errors": [f"Erreur : {str(e)}"]})


from django.shortcuts import render
from django.http import HttpResponseRedirect
from openpyxl import load_workbook
from tablib import Dataset
from .models import Cart
from .resources import CartResource


from django.shortcuts import render
from django.http import HttpResponseRedirect
from openpyxl import load_workbook
from tablib import Dataset
from .models import Cart
from .resources import CartResource


class CartImportView(LoginRequiredMixin , View):
    def get(self, request, *args, **kwargs):
        return render(request, "myapp/cart_import.html")

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return render(request, "myapp/cart_import.html", {"errors": ["Aucun fichier n'a été fourni."]})

        # Vérification du format du fichier
        if not (file.name.endswith('.xls') or file.name.endswith('.xlsx')):
            return render(request, "myapp/cart_import.html", {"errors": ["Seuls les fichiers Excel sont acceptés."]})

        try:
            # Charger le fichier Excel
            workbook = load_workbook(file)
            sheet = workbook.active

            # Vérifier les colonnes attendues
            required_columns = ['nom', 'description']
            detected_columns = [cell.value for cell in sheet[1]]  # Première ligne pour les en-têtes
            missing_columns = [col for col in required_columns if col not in detected_columns]

            if missing_columns:
                return render(request, "myapp/cart_import.html", {
                    "errors": [
                        f"Le fichier Excel est invalide. Colonnes manquantes : {', '.join(missing_columns)}",
                        f"Colonnes détectées : {', '.join(detected_columns or [])}"
                    ]
                })

            # Préparer les données pour l'importation
            resource = CartResource()
            dataset = Dataset()
            dataset.headers = required_columns

            # Lire les données du fichier Excel à partir de la deuxième ligne
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Vérifie que le champ 'nom' est renseigné
                    dataset.append([row[0], row[1]])

            # Simulation de l'importation avec dry_run=True
            result = resource.import_data(dataset, dry_run=True)
            if not result.has_errors():
                # Importation réelle
                resource.import_data(dataset, dry_run=False)
                return HttpResponseRedirect('/carts/')
            else:
                return render(request, "myapp/cart_import.html", {"errors": result.row_errors})

        except Exception as e:
            return render(request, "myapp/cart_import.html", {"errors": [f"Erreur lors de la lecture du fichier Excel : {str(e)}"]})


from django.shortcuts import render, redirect
from tablib import Dataset
from .models import Product, Cart, CartProduct
from datetime import datetime
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View

class CartProductImportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        return render(request, "myapp/cartproduct_import.html")

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return render(request, "myapp/cartproduct_import.html", {"errors": ["Aucun fichier n'a été fourni."]})

        try:
            # Charger les données du fichier Excel
            dataset = Dataset()
            imported_data = dataset.load(file.read(), format='xlsx')

            # Vérifier les colonnes nécessaires
            required_columns = ['produit', 'panier', 'poids', 'date_from', 'date_to']
            detected_columns = dataset.headers
            missing_cols = [col for col in required_columns if col not in detected_columns]

            if missing_cols:
                return render(request, "myapp/cartproduct_import.html", {
                    "errors": [
                        f"Le fichier Excel est invalide. Colonnes manquantes : {', '.join(missing_cols)}",
                        f"Colonnes détectées : {', '.join(detected_columns or [])}"
                    ]
                })

            errors = []  # Liste des erreurs
            for i, row in enumerate(imported_data.dict, start=2):  # Démarre à la 2ème ligne pour ignorer l'en-tête
                try:
                    # Recherche du produit
                    produits = Product.objects.filter(nom=row['produit'])
                    if produits.count() == 1:
                        produit = produits.first()
                    elif produits.count() > 1:
                        errors.append(f"Ligne {i}: Plusieurs produits trouvés pour '{row['produit']}'.")
                        continue
                    else:
                        errors.append(f"Ligne {i}: Produit '{row['produit']}' introuvable.")
                        continue

                    # Recherche du panier
                    panier = Cart.objects.filter(nom=row['panier']).first()
                    if not panier:
                        errors.append(f"Ligne {i}: Panier '{row['panier']}' introuvable.")
                        continue

                    # Vérification et conversion des dates
                    try:
                        date_from = datetime.strptime(row['date_from'], "%Y-%m-%d").date()
                        date_to = datetime.strptime(row['date_to'], "%Y-%m-%d").date()
                    except ValueError:
                        errors.append(f"Ligne {i}: Format de date invalide pour '{row['date_from']}' ou '{row['date_to']}'.")
                        continue

                    # Vérifier que date_to est après date_from
                    if date_to < date_from:
                        errors.append(f"Ligne {i}: La date de fin '{row['date_to']}' est antérieure à la date de début '{row['date_from']}'.")
                        continue

                    # Création ou mise à jour de la relation produit-panier
                    CartProduct.objects.update_or_create(
                        produit=produit,
                        panier=panier,
                        defaults={
                            'poids': row['poids'],
                            'date_from': date_from,
                            'date_to': date_to
                        }
                    )
                except Exception as e:
                    errors.append(f"Ligne {i}: Erreur inattendue : {str(e)}")

            if errors:
                return render(request, "myapp/cartproduct_import.html", {"errors": errors})

            return redirect('cartproduct_list')

        except Exception as e:
            return render(request, "myapp/cartproduct_import.html", {"errors": [f"Erreur lors de l'importation : {str(e)}"]})



from django.views import View
from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import Avg
from datetime import datetime
from .models import Product, ProductPrice, Cart, CartProduct


class INPCCalculateView(LoginRequiredMixin, View):
    """
    Vue pour calculer l'INPC en utilisant les données existantes.
    """

    def calculate_inpc_for_date(self, date):
        """
        Fonction utilitaire pour calculer l'INPC pour une date donnée.
        """
        # Étape 1 : Calcul du prix moyen des produits pour une date donnée
        product_avg_prices = {}
        for product in Product.objects.all():
            avg_price = ProductPrice.objects.filter(
                produit=product,
                date_from__lte=date,
                date_to__gte=date
            ).aggregate(Avg('valeur'))['valeur__avg']

            if avg_price is not None:
                product_avg_prices[product.id] = avg_price

        if not product_avg_prices:
            return 0  # Aucun produit ou prix trouvé

        # Étape 2 : Calcul de l'INPC pour chaque panier
        cart_inpc = {}
        for cart in Cart.objects.all():
            total_weighted_price = 0
            total_weight = 0

            cart_products = CartProduct.objects.filter(
                panier=cart,
                date_from__lte=date,
                date_to__gte=date
            )

            for cart_product in cart_products:
                product_id = cart_product.produit.id
                if product_id in product_avg_prices:
                    total_weighted_price += product_avg_prices[product_id] * cart_product.poids
                    total_weight += cart_product.poids

            # INPC pour le panier
            if total_weight > 0:
                cart_inpc[cart.id] = total_weighted_price / total_weight
            else:
                cart_inpc[cart.id] = 0

        if not cart_inpc:
            return 0  # Aucun panier trouvé

        # Étape 3 : Calcul de l'INPC global
        return sum(cart_inpc.values()) / len(cart_inpc)

    def get(self, request):
        """
        Affiche un formulaire permettant de sélectionner l'année et le mois.
        """
        current_year = datetime.now().year
        return render(request, 'myapp/inpc_form.html', {'current_year': current_year})

    def post(self, request):
        """
        Traite le formulaire soumis par l'utilisateur pour calculer l'INPC.
        """
        try:
            # Récupération des données du formulaire
            month = int(request.POST.get('month'))
            year = int(request.POST.get('year'))
            date = datetime(year, month, 1)

            # Année de base (2019)
            base_year = 2019
            base_date = datetime(base_year, 1, 1)

            # Calcul INPC pour l'année de base
            base_inpc = self.calculate_inpc_for_date(base_date)

            # Calcul INPC pour la période actuelle
            current_inpc = self.calculate_inpc_for_date(date)

            # Normalisation de l'INPC par rapport à l'année de base
            if base_inpc > 0:
                inpc_global = (current_inpc / base_inpc) * 100
            else:
                inpc_global = 0

            # Préparer les données pour le template
            context = {
                'inpc': inpc_global,
                'month': month,
                'year': year,
            }
            return render(request, 'myapp/inpc_result.html', context)

        except Exception as e:
            messages.error(request, f"Erreur lors du calcul de l'INPC : {str(e)}")
            return redirect('calculate_inpc')


from django.shortcuts import render
from django.views import View
from django.db.models import Avg
from datetime import datetime
from .models import ProductPrice, CartProduct, Product, Cart

class InpcListView(LoginRequiredMixin , View):
    template_name = "myapp/inpc_list.html"

    def get(self, request, *args, **kwargs):
        try:
            # Obtenir l'année et le mois actuels
            year = request.GET.get('year', datetime.now().year)
            month = request.GET.get('month', datetime.now().month)

            # Convertir en entiers
            year = int(year)
            month = int(month)

            # Année de base (2019)
            base_year = 2019
            base_date = datetime(base_year, 1, 1)

            # Calculer l'INPC pour la période actuelle
            date = datetime(year, month, 1)
            base_inpc = self.calculate_inpc(base_date)
            current_inpc = self.calculate_inpc(date)

            # Normaliser l'INPC par rapport à l'année de base
            inpc_global = (current_inpc / base_inpc * 100) if base_inpc > 0 else 0

            context = {
                'inpc_global': inpc_global,
                'current_year': year,
                'current_month': month,
            }
            return render(request, self.template_name, context)

        except Exception as e:
            context = {'error': f"Erreur lors du calcul de l'INPC : {str(e)}"}
            return render(request, self.template_name, context)

    def calculate_inpc(self, date):
        # Calcul du prix moyen pour chaque produit
        product_avg_prices = {}
        for product in Product.objects.all():
            avg_price = ProductPrice.objects.filter(
                produit=product,
                date_from__lte=date,
                date_to__gte=date
            ).aggregate(Avg('valeur'))['valeur__avg']

            if avg_price is not None:
                product_avg_prices[product.id] = avg_price

        if not product_avg_prices:
            return 0

        # Calcul de l'INPC pour chaque panier
        cart_inpc = {}
        for cart in Cart.objects.all():
            total_weighted_price = 0
            total_weight = 0

            cart_products = CartProduct.objects.filter(
                panier=cart,
                date_from__lte=date,
                date_to__gte=date
            )

            for cart_product in cart_products:
                product_id = cart_product.produit.id
                if product_id in product_avg_prices:
                    total_weighted_price += product_avg_prices[product_id] * cart_product.poids
                    total_weight += cart_product.poids

            cart_inpc[cart.id] = total_weighted_price / total_weight if total_weight > 0 else 0

        # Calcul INPC global
        return sum(cart_inpc.values()) / len(cart_inpc) if cart_inpc else 0
    
    
#Charts
from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import Avg
from .models import ProductPrice, Product
@login_required

def dashboard_view(request):
    products = Product.objects.all()
    return render(request, "myapp/dashboard.html", {"products": products})
@login_required

def get_product_price_data(request):
    product_id = request.GET.get("product_id")
    aggregation = request.GET.get("aggregation", "monthly")  # Valeur par défaut = mensuelle

    if not product_id:
        return JsonResponse({"dates": [], "prices": [], "product_name": "Sélectionner un produit"})

    prices = ProductPrice.objects.filter(produit_id=product_id)

    if aggregation == "monthly":
        prices = prices.values("date_from__year", "date_from__month").annotate(avg_price=Avg("valeur")).order_by("date_from__year", "date_from__month")
        data = {
            "dates": [f"{p['date_from__year']}-{p['date_from__month']:02d}" for p in prices],
            "prices": [p["avg_price"] for p in prices],
        }
    elif aggregation == "yearly":
        prices = prices.values("date_from__year").annotate(avg_price=Avg("valeur")).order_by("date_from__year")
        data = {
            "dates": [str(p["date_from__year"]) for p in prices],
            "prices": [p["avg_price"] for p in prices],
        }
    else:
        return JsonResponse({"dates": [], "prices": [], "product_name": "Aucune donnée"})

    data["product_name"] = Product.objects.get(id=product_id).nom if prices else "Produit inconnu"
    return JsonResponse(data)



from django.http import JsonResponse
from django.db.models import Avg
from .models import ProductPrice, Product
from django.http import JsonResponse
from django.db.models import Avg
from .models import ProductPrice, Product
@login_required

def bar_chart_product_prices(request):
    """
    Vue qui génère les données pour un Bar Chart montrant l'évolution des prix moyens par mois.
    """
    product_id = request.GET.get("product_id")

    if not product_id:
        return JsonResponse({"dates": [], "prices": [], "product_name": "Sélectionner un produit"})

    # Filtrer les prix du produit sélectionné
    prices = (
        ProductPrice.objects.filter(produit_id=product_id)
        .values("date_from__year", "date_from__month")
        .annotate(avg_price=Avg("valeur"))
        .order_by("date_from__year", "date_from__month")
    )

    data = {
        "dates": [f"{p['date_from__year']}-{p['date_from__month']:02d}" for p in prices],
        "prices": [p["avg_price"] for p in prices],
        "product_name": Product.objects.get(id=product_id).nom if prices else "Produit inconnu",
    }

    return JsonResponse(data)

from django.http import JsonResponse
from django.db.models import Avg
from .models import ProductPrice

@login_required
def bar_chart_all_products_prices(request):
    """
    Vue qui génère les données pour un Bar Chart montrant l'évolution des prix moyens mensuels pour tous les produits.
    """
    from django.db.models import Avg

    prices = (
        ProductPrice.objects
        .values("date_from__year", "date_from__month")
        .annotate(avg_price=Avg("valeur"))
        .order_by("date_from__year", "date_from__month")
    )

    data = {
        "dates": [f"{p['date_from__year']}-{p['date_from__month']:02d}" for p in prices],
        "prices": [p["avg_price"] for p in prices],
    }

    return JsonResponse(data)

from django.http import JsonResponse
from django.db.models import Count
from .models import Product, ProductType  # Assurez-vous que ProductType est bien défini
from django.http import JsonResponse
from django.db.models import Count
from .models import Product, ProductType

@login_required
def pie_product_categories(request):
    """
    Génère les données pour un Pie Chart montrant la répartition des produits par catégorie.
    """
    categories = ProductType.objects.annotate(count=Count("produits"))
    
    data = {
        "labels": [category.nom for category in categories],
        "values": [category.count for category in categories],
    }

    return JsonResponse(data)

#################################################################33
import numpy as np
import json
from datetime import datetime, timedelta
from django.http import JsonResponse
from django.db.models import Avg
from .models import ProductPrice, CartProduct, Inpc
@login_required

def get_inpc_chart_data(request):
    # Définition de la période de référence (ex : janvier 2023)
    reference_month = datetime(2023, 1, 1)
    
    # Récupérer les prix moyens par mois
    prices_by_month = (
        ProductPrice.objects
        .values('date_from__year', 'date_from__month', 'produit')
        .annotate(avg_price=Avg('valeur'))
        .order_by('date_from__year', 'date_from__month')
    )
    
    # Construire un dictionnaire des prix par produit et par mois
    prices_dict = {}
    for p in prices_by_month:
        year_month = (p['date_from__year'], p['date_from__month'])
        if year_month not in prices_dict:
            prices_dict[year_month] = {}
        prices_dict[year_month][p['produit']] = p['avg_price']
    
    # Récupérer les poids des produits
    product_weights = {cp.produit.id: cp.poids for cp in CartProduct.objects.all()}
    
    # Trouver les prix de référence
    ref_prices = prices_dict.get((reference_month.year, reference_month.month), {})
    
    # Calculer l'INPC pour chaque mois
    inpc_values = []
    dates = []
    for (year, month), price_data in prices_dict.items():
        numerator = sum(price_data.get(prod, 0) * product_weights.get(prod, 1) for prod in price_data)
        denominator = sum(ref_prices.get(prod, 1) * product_weights.get(prod, 1) for prod in ref_prices)
        if denominator > 0:
            inpc = (numerator / denominator) * 100
            inpc_values.append(inpc)
            dates.append(f"{year}-{month:02d}")
    
    # Sauvegarder les valeurs INPC dans la base de données
    for date_str, value in zip(dates, inpc_values):
        Inpc.objects.update_or_create(mois=datetime.strptime(date_str, "%Y-%m"), defaults={'valeur': value})
    
    # Préparer les données pour Chart.js
    chart_data = {
        'labels': dates,
        'datasets': [{
            'label': "Évolution de l'INPC",
            'data': inpc_values,
            'borderColor': 'rgba(75, 192, 192, 1)',
            'fill': False,
            'tension': 0.4
        }]
    }
    
    return JsonResponse(chart_data)

from django.shortcuts import render
from django.db.models import F
from .models import Inpc2
import datetime
@login_required

def calculate_inpc_2(request):
    selected_year = request.GET.get('year', datetime.datetime.today().year)
    selected_month = request.GET.get('month', datetime.datetime.today().month)

    try:
        selected_date = datetime.date(int(selected_year), int(selected_month), 1)
    except ValueError:
        selected_date = None

    inpc_value = None
    message = "Aucune donnée INPC trouvée pour cette année."

    if selected_date:
        # Vérifier si des données existent pour cette année
        year_data_exists = Inpc2.objects.filter(mois__year=selected_year).exists()

        if year_data_exists:
            # Récupérer la valeur INPC pour le mois sélectionné
            inpc_entry = Inpc2.objects.filter(mois=selected_date).first()

            if inpc_entry:
                inpc_value = inpc_entry.valeur

                # Récupérer la valeur INPC du mois précédent
                previous_month = selected_date - datetime.timedelta(days=1)
                previous_month = previous_month.replace(day=1)
                previous_inpc_entry = Inpc2.objects.filter(mois=previous_month).first()

                if previous_inpc_entry:
                    previous_value = previous_inpc_entry.valeur
                    variation = ((inpc_value - previous_value) / previous_value) * 100 if previous_value else 0

                    if variation > 0:
                        message = f"Le mois de {selected_date.strftime('%B %Y')} a subi une augmentation de {variation:.2f}%."
                    elif variation < 0:
                        message = f"Le mois de {selected_date.strftime('%B %Y')} a subi une diminution de {abs(variation):.2f}%."
                    else:
                        message = f"Le mois de {selected_date.strftime('%B %Y')} est stable."
                else:
                    message = f"Aucune donnée INPC pour le mois précédent ({previous_month.strftime('%B %Y')})."
            else:
                message = f"Aucune donnée INPC trouvée pour {selected_date.strftime('%B %Y')}."
        else:
            message = f"Aucune donnée INPC trouvée pour l'année {selected_year}."

    return render(request, 'calculate_inpc_2.html', {
        'selected_year': selected_year,
        'selected_month': selected_month,
        'inpc_value': inpc_value,
        'message': message
    })

from django.http import JsonResponse
from .models import Inpc
@login_required

def get_inpc_chart_data(request):
    """Récupère les données INPC pour le graphique"""
    inpc_entries = Inpc.objects.order_by('mois')

    # Extraire les dates et valeurs INPC
    dates = [entry.mois.strftime("%Y-%m") for entry in inpc_entries]
    values = [entry.valeur for entry in inpc_entries]

    # Préparer les données pour Chart.js
    chart_data = {
        'labels': dates,
        'datasets': [{
            'label': "Évolution de l'INPC",
            'data': values,
            'borderColor': 'rgba(75, 192, 192, 1)',
            'backgroundColor': 'rgba(75, 192, 192, 0.2)',
            'fill': True,
            'tension': 0.4
        }]
    }

    return JsonResponse(chart_data)

from django.shortcuts import render
from .models import Inpc2
@login_required

def list_inpc_2(request):
    # Récupérer toutes les valeurs INPC triées par date (du plus récent au plus ancien)
    inpc_data = Inpc2.objects.all().order_by('-mois')

    return render(request, 'myapp/list_inpc_2.html', {
        'inpc_data': inpc_data
    })


from .models import Inpc2
@login_required
def list_inpc_2(request):
    # Récupération du filtre de date si fourni
    date_filter = request.GET.get('date_filter', '')

    # Filtrage des données INPC
    if date_filter:
        year, month = map(int, date_filter.split('-'))
        inpc_data = Inpc2.objects.filter(mois__year=year, mois__month=month).order_by('-mois')
    else:
        inpc_data = Inpc2.objects.all().order_by('-mois')  # Trier par date descendante

    # PAGINATION
    paginator = Paginator(inpc_data, 10)  # 10 résultats par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'myapp/list_inpc_2.html', {
        'page_obj': page_obj,
        'is_paginated': paginator.num_pages > 1,
        'request': request
    })
